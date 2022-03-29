import numpy as np
from openpyxl import load_workbook
u=[]
k=[]
iu = load_workbook(filename='iu.xlsx')
iu1 = load_workbook(filename='iu1.xlsx')
sheet = iu.active
sheet1=iu1.active
pt = sheet.max_row-4
tp = sheet.max_column
for i in range(3,pt):
    u.append(sheet1.cell(i,2).value)
    k.append(sheet1.cell(i, 1).value)
gh=[]
print(u)
k1=[]
for j in range(len(u)):
    for i in range(2,tp):
        if u[j] == sheet.cell(10, i).value:
            for y in range(3,pt):
                gh.append(sheet.cell(y,i).value)
                if j==len(u)-1:
                    k1.append(sheet.cell(y,tp).value)
A = (np.array(gh,dtype=float).reshape((pt-3,len(u)))).swapaxes(0,1)
print('\nD')
print(A)
print('\nзапасы ресурсов b= ',k1)
print('\nCb',k)
D = np.array(np.matrix(A, float).getI(), float)#D^-1
print('\nD^(-1)',)
print(D)
j = np.array(np.dot(k, D))
print('\nдефицитные ресурсы  y^* = D^(-1)*Cb = ',j)
print('g(min)=b*(y^*)=',np.dot(k1, j))