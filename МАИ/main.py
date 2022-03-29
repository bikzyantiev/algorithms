from openpyxl import load_workbook

b = load_workbook(filename='iu3.xlsx')
c = load_workbook(filename='iu4.xlsx')
sheet1 = b.active
my_sheet = c.active
pt = sheet1.max_row-4
V = 1
F = 1
y = [(pt)*[0]]*(pt+1)
print(y)
c.active = 0
my_sheet = c.active
for i in range(10, 25):
    j = 1
    if my_sheet.cell(i, j).value == pt:
        pp = my_sheet.cell(i, j+1).value

def o(sheet):
   for i in range(1, pt + 2):
       for j in range(1, pt + 2):
           if j == 1 and i != 1:
               sheet.cell(i, j).value = i - 1
           if i == 1 and j != 1:
               sheet.cell(i, j).value = j - 1
   c.save("iu4.xlsx")


o(my_sheet)
# for i in range(0, pt):
#     c.create_sheet(index=i + 1, title="k" + str(i + 1))
#     c.save("D:/тпр/МАИ/iu4.xlsx")
for i in range(1, pt + 1):
    c.active = i
    sheet = c.active
    o(sheet)
def u(sheet,i):
    for j in range(3, pt + 3):
        for z in range(3,pt +3):
            if (int(sheet1.cell(j,i).value - sheet1.cell(z,i).value)) >= 12:
                sheet.cell(j-1, z-1).value = 1/9
                sheet.cell(z-1, j-1).value = 9
            elif 12 > (int(sheet1.cell(j,i).value - sheet1.cell(z,i).value)) >= 11:
                sheet.cell(j-1, z-1).value = 1/7
                sheet.cell(z-1, j-1).value = 7
            elif 8 <= int(sheet1.cell(j,i).value - sheet1.cell(z,i).value) < 11:
                sheet.cell(j-1, z-1).value = 1/5
                sheet.cell(z-1, j-1).value = 5
            elif 0 < int(sheet1.cell(j,i).value - sheet1.cell(z,i).value) < 8:
                sheet.cell(j-1, z-1).value = 1/3
                sheet.cell(z-1, j-1).value = 3
            elif sheet1.cell(j,i).value == sheet1.cell(z,i).value:
                sheet.cell(j-1, z-1).value = 1
                sheet.cell(z-1, j-1).value = 1
for j in range(2, pt + 2):
    for k in range(2, pt + 2):
        if (sheet1.cell(pt+3, j).value - sheet1.cell(pt+3, k).value) >= 11:
            my_sheet.cell(k, j).value = 9
            my_sheet.cell(j, k).value = 1 / 9
        elif 11>(sheet1.cell(pt+3, j).value - sheet1.cell(pt+3, k).value) >= 8:
            my_sheet.cell(k, j).value = 7
            my_sheet.cell(j, k).value = 1 / 7
        elif 6 <= (sheet1.cell(pt+3, j).value - sheet1.cell(pt+3, k).value) < 8:
            my_sheet.cell(k, j).value = 5
            my_sheet.cell(j, k).value = 1 / 5
        elif 0 < (sheet1.cell(pt+3, j).value - sheet1.cell(pt+3, k).value) < 6:
            my_sheet.cell(k, j).value = 3
            my_sheet.cell(j, k).value = 1 / 3
        elif sheet1.cell(pt+3, j).value == sheet1.cell(pt+3, k).value:
            my_sheet.cell(k, j).value = 1
            my_sheet.cell(j, k).value = 1
for i in range(1, pt + 1):
    c.active = i
    sheet = c.active
    u(sheet,i+1)
c.save("iu4.xlsx")
def g(sheet,e):
    x = []
    x1 = []
    x2 = []
    x3 = []
    sheet.cell(1, pt + 2).value="V"
    for i in range(2, pt + 2):
        V = 1
        for j in range(2, pt + 2):
            V *= float(sheet.cell(i, j).value)
        sheet.cell(i, pt + 2).value = round(V ** (1 / pt), 4)
        x.append(round(V ** (1 / pt), 3))
    ZV = sum(x)
    sheet.cell(pt+2, pt + 2).value=ZV
    sheet.cell(pt + 2, pt + 1).value = "sumV"
    sheet.cell(1, pt + 3).value = "W"
    for i in range(0, len(x)):
        W = x[i] / ZV
        sheet.cell(i+2, pt + 3).value = round(W, 3)
        x1.append(round(W, 3))
        y[e] = x1
    sheet.cell(1, pt + 4).value = "S"
    sheet.cell(1, pt + 5).value = "P"
    sheet.cell(pt + 2, pt + 4).value = "L"
    for i in range(2, pt + 2):
        S = 0
        for j in range(2, pt + 2):
            S += sheet.cell(j, i).value
        sheet.cell(i, pt + 4).value = round(S, 3)
        x2.append(round(S, 4))
    for i in range(0, len(x1)):
        sheet.cell(i+2, pt + 5).value = round(x1[i] * x2[i], 3)
        x3.append(round(x1[i] * x2[i], 4))
    l = round(sum(x3), 3)
    sheet.cell(pt+2, pt + 5).value = l
    IS = round(((l - pt) / (pt - 1)), 3)
    print("IS",IS)
    OS = IS / pp
    print("OS",OS)
    c.save("iu4.xlsx")

e=0
for i in range(0, pt + 1):
    c.active = i
    sheet = c.active
    print("\n", i, "\n")
    g(sheet,e)
    e += 1
print('\n')
for i in range(pt+1):

    print(y[i],'\n')
print('\n')
po=0
op=[]
e1=1
for j in range(0,pt):
    for i in range(0,pt):
        po+=y[0][i]*y[i+1][j]
    op.append(po)
    po=0
print(op)
print(max(op))