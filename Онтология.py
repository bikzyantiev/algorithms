from tkinter import *
from tkinter import filedialog as fd
from tkinter import ttk, messagebox
import tkinter as tk
from tkinter.ttk import Combobox
from openpyxl import load_workbook
import openpyxl
import os
from os import walk
from os.path import join

from tkinter import Entry

global sheet, w3, w4, c1, len_row, \
    entry_slot, file_name0, p, w6, \
    k1, l, w9, w10, k3,w11, entry5, lbox, w15

k1 = 2
l = 0
p = 0
file_name0 = os.getcwd()
len_row = 1
file_name8=os.getcwd()

# ----------------------------ctrl+c,ctrl+v,ctrl+x----------------------------
def _onKeyRelease(event):
    ctrl = (event.state & 0x4) != 0
    if event.keycode == 88 and ctrl and event.keysym.lower() != "x":
        event.widget.event_generate("<<Cut>>")

    if event.keycode == 86 and ctrl and event.keysym.lower() != "v":
        event.widget.event_generate("<<Paste>>")

    if event.keycode == 67 and ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")


# ----------------------------окно сохранения---------------------------------

def extract_text():
    s = ''
    file_name = fd.asksaveasfilename(
        filetypes=(("папка с проектом", "*.flk"),
                   ("All files", "*.*")))
    global file_name1
    file_name1 = file_name
    while file_name[-1] != "/":
        s += file_name[-1]
        file_name = file_name[:-1]
    s = s[::-1]
    os.chdir(str(file_name))
    os.mkdir(s)
    os.chdir(str(file_name1))
    return(s)

# ------------------------поиск пути к выбранному файлу-------------------------
def find(row, filepath):
    d = int(row)
    global way
    way = []
    k = filepath
    file_name2 = file_name1

    while d > 1:
        for j in range(2, len1(d)):
            if sheet.cell(d, j).value == k:
                way.append(sheet.cell(d, j).value)
                k = sheet.cell(d, j + 1).value
                if d == 2:
                    way.append(sheet.cell(d, j + 1).value)
                break
        d -= 1

    for d in range(len(way) - 1, -1, -1):
        if os.path.isdir(file_name2 + "/" + way[d]) == 1:
            file_name2 = file_name2 + "/" + str(way[d])
            os.chdir(str(file_name2))
        else:
            os.mkdir(way[d])
            file_name2 = file_name2 + "/" + str(way[d])
            os.chdir(str(file_name2))
    return (file_name2)


# -----------------------------------считает длину строки----------------------------------------
def len1(f):
    flag = 0
    i = 1
    os.chdir(file_name1)
    bw = load_workbook(filename='буфер.xlsx')
    sheet = bw.active
    while flag == 0:
        if sheet.cell(f, i + 1).value is None and sheet.cell(f, i).value is not None:
            flag = 1
        else:
            i += 1
    return i


# --------------------------фунция поддверждения закрытия программы-------------------------------
def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        window.destroy()


# ----------------------------------------функции кнопок-------------------------------------------
def button1Callback():
    global w1, choices1, choices2, w3, w4, w, w6, w9,w11, c, file_name0
    os.chdir(str(file_name1))
    c1 = []
    if message1.get() == "" or w.get() == "":
        messagebox.showinfo('Ошибка', 'Введите все поля')
    else:
        global k1, l
        choices1.append(entry1.get())
        sheet.cell(1, k1).value = entry1.get()
        sheet.cell(1, k1 + 1).value = "*"
        entry1.delete(0, END)
        w.delete(0, END)
        w1.destroy()
        w1 = Combobox(tab1, values=choices1)
        w1.place(x=240, y=70, height=25, width=106)
        os.chdir(str(file_name1))
        b.save("буфер.xlsx")
        if len(c) == 0:
            c.append(2)
        if l == 0:
            w3 = Combobox(tab1, values=c)
            w3.place(relx=0.19, y=70, height=25, relwidth=0.15)
        for i in range(len(c)):
            c1.append(c[i])
        if c1[0] != 1:
            c1.insert(0, 1)
        w4.destroy()
        w4 = Combobox(tab2, values=c1)
        w4.place(x=240, y=24, height=25, width=106)
        w9.destroy()
        w9 = Combobox(tab4, values=c1)
        w9.place(x=0, y=30, height=25, width=106)
        w6.destroy()
        w6 = Combobox(tab3, values=c1)
        w6.place(x=0, y=30, height=25, width=106)
        w11.destroy()
        w11 = Combobox(tab5, values=c1)
        w11.place(x=0, y=30, height=25, width=106)
        k1 += 2
        l += 1


def button2Callback():
    c = []
    global w3, f, len_row, w1, w4, w6, c1, w9,w11
    f = int(w3.get())
    sheet.cell(f, 1).value = f
    i = 1
    flag = 0
    while flag == 0:
        if sheet.cell(f, i + 1).value is None and sheet.cell(f, i).value is not None:
            len_row = i
            flag = 1
        else:
            i += 1
    sheet.cell(f, len_row + 1).value = entry2.get()
    sheet.cell(f, len_row + 2).value = w1.get()
    sheet.cell(f + 1, 1).value = f + 1
    for i in range(2, sheet.max_row + 1):
        c.append(i)
    w3 = Combobox(tab1, values=c)
    w3.place(x=130, y=70, height=25, width=106)
    c1 = c
    if c1[0] != 1:
        c1.insert(0, 1)
    w4.destroy()
    w4 = Combobox(tab2, values=c1)
    w4.place(x=240, y=24, height=25, width=106)
    w9.destroy()
    w9 = Combobox(tab4, values=c1)
    w9.place(x=0, y=30, height=25, width=106)
    w6.destroy()
    w6 = Combobox(tab3, values=c1)
    w6.place(x=0, y=30, height=25, width=106)
    w11.destroy()
    w11 = Combobox(tab5, values=c1)
    w11.place(x=0, y=30, height=25, width=106)
    entry2.delete(0, END)
    w1.delete(0, END)
    w3.delete(0, END)
    os.chdir(str(file_name1))
    b.save("буфер.xlsx")


def button3Callback():
    global w3, f, len_row, choices3, w1
    f = int(w3.get())
    choices3 = choices1
    if f != 2:
        flag = 0
        i = 1
        while flag == 0:
            if sheet.cell(f - 1, i + 1).value is None and sheet.cell(f - 1, i).value is not None:
                len_row = i
                flag = 1
            else:
                i += 1
        choices3 = []
        for i in range(2, len_row + 1, 2):
            choices3.append(sheet.cell(f - 1, i).value)
    w1.destroy()
    w1 = Combobox(tab1, values=choices3)
    w1.place(x=240, y=70, height=25, width=106)
    os.chdir(str(file_name1))
    b.save("буфер.xlsx")


def button4Callback():
    global w4, choices_tab2, w5, w3
    choices_tab2 = []
    f = int(w4.get())
    len_row = len1(f)
    choices5=[]
    i=1
    flag=0
    while flag == 0:
        if sheet.cell(f, i + 1).value is None and sheet.cell(f, i).value is not None:
            len_row = i
            flag = 1
        else:
            i += 1
    choices5 = []
    for i in range(2, len_row + 1, 2):
        choices5.append(sheet.cell(f, i).value)
    w5.destroy()
    w5 = Combobox(tab2, values=choices5)
    w5.place(x=350, y=24, height=25, width=106)

def button51Callback(w42):
    global w51
    f = int(w42.get())

    flag = 0
    i = 1
    while flag == 0:
        if sheet.cell(f, i + 1).value is None and sheet.cell(f, i).value is not None:
            len_row = i
            flag = 1
        else:
            i += 1
    choices6 = []
    for i in range(2, len_row + 1, 2):
        choices6.append(sheet.cell(f, i).value)
    w51.destroy()
    w51 = Combobox(tab2, values=choices6)
    w51.place(x=120, y=60, height=25, width=106)
def button511Callback():
    print(w51.get())
    if sheet1.max_row == 0:
        sheet1.cell(2, 1).value = 'inst_' + entry_slot.get() + "_" + w42.get() + "_" + w51.get()
    else:
        sheet1.cell(sheet1.max_row + 1, 1).value = 'inst_' + entry_slot.get() + "_" + w42.get() + "_" + w51.get()
    bw.save('таблица.xlsx')

def button5Callback():
    global w5, w4, w3, entry_slot, w6, way, file_name2, w42, w51 ,sheet1,bw
    i1 = int(w4.get())
    k = w5.get()

    os.chdir(str(file_name1))
    bw1 = load_workbook(filename='буфер.xlsx')
    sheet8 = bw1.active
    bw1.save('буфер.xlsx')
    global obj
    obj = []
    obj.append(w5.get())
    i2 = w4.get()

    file_name2 = find(i1, k)
    if (os.path.isdir(file_name2 + "/" + k) == 0 and int(i2) == 1) or (os.path.isdir(file_name2) == 0 and int(i2) != 1):
        os.mkdir(file_name2 + "/" + k)
    if int(w4.get()) == 1:
        file_name5 = file_name2 + f'/{k}'
    else:
        file_name5 = file_name2
    os.chdir(file_name5)
    if os.path.exists(file_name5 + '/таблица.xlsx') == 1:
        bw = load_workbook(filename='таблица.xlsx')
        sheet1 = bw.active
        bw.save('таблица.xlsx')
    else:
        bw = openpyxl.Workbook()
        bw.save('таблица.xlsx')
        bw = load_workbook(filename='таблица.xlsx')
        sheet1 = bw.active
    bw.save('таблица.xlsx')
    os.chdir(file_name2)
    if w41.get()!='instance':
        if sheet1.max_row == 0:
            sheet1.cell(2, 1).value = entry_slot.get()
        else:
            sheet1.cell(sheet1.max_row + 1, 1).value = entry_slot.get()
    else:
        c = []
        c1 = []
        for i in range(2, sheet.max_row + 1):
            c.append(i)
        print(c)
        for i in range(len(c)):
            c1.append(c[i])
        if c1[0] != 1:
            c1.insert(0, 1)
        w42 = Combobox(tab2, values=c1)
        w42.place(x=0, y=60, height=25, width=106)
        choices6=[]
        w51 = Combobox(tab2, values=choices6)
        w51.place(x=120, y=60, height=25, width=106)
        butt = Button(tab2, text='ok', bg='lightgreen', command=lambda: button51Callback(w42))
        butt.place(x=80, y=90, height=25, relwidth=0.05)
        butt2 = Button(tab2, text='ok', bg='lightgreen', command=button511Callback)
        butt2.place(x=150, y=90, height=25, relwidth=0.05)


###################################################################################################################################################################
    bw.save(file_name5 + '/таблица.xlsx')
    file_name2 = file_name5

    for i3 in range(int(i2), sheet8.max_row + 1):
        uk = i3
        obj1 = []
        l = []
        l1 = []


        for j1 in range(len(obj)):
            pj = len1(uk + 1)
            for z in range(3, pj + 1, 2):
                if obj[j1] == sheet8.cell(i3 + 1, z).value:
                    obj1.append(sheet8.cell(i3 + 1, z - 1).value)



        for j in range(len(obj)):
            bw = load_workbook(file_name2 + '/таблица.xlsx')
            sheet1 = bw.active
            for k in range(2, sheet1.max_row + 1):
                l.append(sheet1.cell(k, 1).value)
            find(uk, obj[j])

            if i3 != int(i2):
                if os.path.exists(os.getcwd() + '/таблица.xlsx') == 1:
                    w = load_workbook(os.getcwd() + '/таблица.xlsx')
                    sheet2 = w.active

                    for k1 in range(2, sheet2.max_row + 1):
                        l1.append(sheet2.cell(k1, 1).value)

                    for k1 in range(len(l)):
                        if l1.count(l[k1]) != 1:
                            l1.append(l[k1])

                    for k1 in range(len(l1)):
                        sheet2.cell(k1 + 2, 1).value = l1[k1]

                    w.save(os.getcwd() + '/таблица.xlsx')
                else:
                    w = openpyxl.Workbook()
                    w.save(os.getcwd() + '/таблица.xlsx')
                    w = load_workbook(os.getcwd() + '/таблица.xlsx')
                    sheet2 = w.active
                    for k in range(len(l)):
                        sheet2.cell(k + 2, 1).value = l[k]
                w.save(os.getcwd() + '/таблица.xlsx')

            l = []
            l1 = []
        obj = obj1



def button6Callback():
    global w6, file_name0, w7
    os.chdir(str(file_name0))
    choices_tab3 = []
    f1 = int(w6.get())
    len_row = len1(f1)
    for i in range(2, len_row + 1, 2):
        choices_tab3.append(sheet.cell(f1, i).value)
    w7.destroy()
    w7 = Combobox(tab3, values=choices_tab3)
    w7.place(x=110, y=30, height=25, width=106)


def button7Callback():
    global w8
    file_name2 = find(w6.get(), w7.get())
    if int(w6.get()) == 1:
        file_name2 = file_name2 + "/" + w7.get()
    g = []
    b = load_workbook(file_name2 + '/таблица.xlsx')
    sheett = b.active
    for i in range(2, sheett.max_row + 1):
        g.append(sheett.cell(i, 1).value)
    w8.destroy()
    w8 = Combobox(tab3, values=g)
    w8.place(x=220, y=30, height=25, width=106)


def button8Callback():
    for dir, subdir, files in walk(find(w6.get(), w7.get())):
        for file in files:
            d = load_workbook(join(dir, file))
            sheet0 = d.active
            k = sheet0.cell(2, 1).value
            for i in range(2, sheet0.max_row + 1):
                if w8.get() == sheet0.cell(i, 1).value:
                    sheet0.cell(2, 1).value = sheet0.cell(i, 1).value
                    sheet0.cell(i, 1).value = k
            d.save(join(dir, file))


def button9Callback():
    global w9, file_name0, w10
    os.chdir(str(file_name0))
    choices_tab3 = []
    f1 = int(w9.get())
    len_row = len1(f1)
    for i in range(2, len_row + 1, 2):
        choices_tab3.append(sheet.cell(f1, i).value)
    w10.destroy()
    w10 = Combobox(tab4, values=choices_tab3)
    w10.place(x=110, y=30, height=25, width=106)

def button101Callback(g,i):
    globals()['swww' + str(i)]+=g.get()+','
    print(globals()['swww' + str(i)])
def button10Callback():
    global w9, w10, file_name6
    file_name6 = find(w9.get(), w10.get())
    if w9.get() == "1":
        file_name6 = file_name6 + "/" + w10.get()
    k = load_workbook(file_name6 + "/таблица.xlsx")
    sheet4 = k.active
    k2 = 120

    for i in range(2, sheet4.max_row + 1):
        globals()['label'+str(i)] = Label(tab4, text=f"{sheet4.cell(i, 1).value}", font=("Arial Bold", 8))
        globals()['label'+str(i)].place(x=0, y=k2 - 20)
        if sheet4.cell(i,1).value.startswith("inst")==1:
            gi=sheet4.cell(sheet4.max_row, 1).value.split('_')
            f=load_workbook(find(gi[len(gi)-2],gi[len(gi)-1])+"/таблица.xlsx")
            sheet5 = f.active
            cj=[]
            for j in  range(2,sheet5.max_column+1):
                cj.append(sheet5.cell(2,j).value)
            globals()['www' + str(i)] = Combobox(tab4,values=cj)
            globals()['www' + str(i)].place(x=0, y=k2)
            globals()['swww' + str(i)]=''
            globals()['buttonwww' + str(i)] = Button(tab4, text='создать класс', bg='lightgreen', command=lambda: button101Callback(globals()['www' + str(i)],i))
            globals()['buttonwww' + str(i)].place(x=160, y=k2)
        else:
            globals()['entry' + str(i)] = Entry(tab4, textvariable=i)
            globals()['entry' + str(i)].place(x=0, y=k2)
        k2 += 50


def button11Callback():
    k = load_workbook(file_name6 + "/таблица.xlsx")
    sheet4 = k.active
    t = sheet4.max_column + 1
    for i in range(2, sheet4.max_row + 1):
        if sheet4.cell(i,1).value.startswith("inst")==1:
            sheet4.cell(i, t).value = globals()['swww' + str(i)]
        else:
            sheet4.cell(i, t).value = globals()['entry' + str(i)].get()
    k.save(file_name6 + "/таблица.xlsx")
    for i in range(2, sheet4.max_row + 1):
        if sheet4.cell(i, 1).value.startswith("inst") == 1:
            globals()['www' + str(i)].delete(0, END)
            globals()['www' + str(i)].destroy()
            globals()['buttonwww' + str(i)].destroy()
        else:
            globals()['entry' + str(i)].delete(0, END)
            globals()['entry' + str(i)].destroy()
        globals()['label' + str(i)].destroy()


def button12Callback():
    global w11,file_name0,w12
    os.chdir(str(file_name0))
    f1 = int(w11.get())
    len_row = len1(f1)
    choices_tab3=[]
    for i in range(2, len_row + 1, 2):
        choices_tab3.append(sheet.cell(f1, i).value)
    w12.destroy()
    w12 = Combobox(tab5, values=choices_tab3)
    w12.place(x=110, y=30, height=25, width=106)

def button13Callback():
    global w13
    file_name2 = find(w11.get(), w12.get())
    if int(w11.get()) == 1:
        file_name2 = file_name2 + "/" + w12.get()
    g = []
    b1 = load_workbook(file_name2 + '/таблица.xlsx')
    sheett1 = b1.active
    for i in range(2, sheett1.max_row + 1):
        g.append(sheett1.cell(i, 1).value)
    w13.destroy()
    w13 = Combobox(tab5, values=g)
    w13.place(x=220, y=30, height=25, width=106)
def button14Callback():
    k=0
    global entry5,lbox,w15,sheet21
    lbox.delete(0, END)
    if os.path.exists(file_name1 + '/буфер1.xlsx') == 1:
        os.remove(file_name1 + '/буфер1.xlsx')
        w1 = openpyxl.Workbook()
        w1.save(file_name1 + '/буфер1.xlsx')
        w1 = load_workbook(file_name1 + '/буфер1.xlsx')
        sheet21 = w1.active
    else:
        w1 = openpyxl.Workbook()
        w1.save(file_name1 + '/буфер1.xlsx')
        w1 = load_workbook(file_name1 + '/буфер1.xlsx')
        sheet21= w1.active
    file_name2 = find(w11.get(), w12.get())
    if int(w11.get()) == 1:
        file_name2 = file_name2 + "/" + w12.get()

    for dir, subdir, files in walk(file_name2):
        for file in files:
            d = load_workbook(join(dir, file))
            sheet0 = d.active
            for i in range(2,sheet0.max_row+1):
                if sheet0.cell(i,1).value==w13.get():
                    k=i
            if w14.get()=="Содержит":
                for i in range(2,sheet0.max_column+1):
                    if (entry5.get()).lower() in (sheet0.cell(k,i).value).lower():
                        p1=sheet21.max_column + 1
                        p2=sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2,sheet0.max_row+1):
                            sheet21.cell(l, p1).value=sheet0.cell(l,i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')


            elif w14.get()=="Не содержит" :
                for i in range(2, sheet0.max_column + 1):
                    if (entry5.get()).lower() not in (sheet0.cell(k,i).value).lower():
                        p1 = sheet21.max_column + 1
                        p2 = sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2, sheet0.max_row + 1):
                            sheet21.cell(l, p1).value = sheet0.cell(l, i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')

            elif w14.get() == "Является":
                for i in range(2, sheet0.max_column + 1):
                    if (entry5.get()).lower() == (sheet0.cell(k, i).value).lower():
                        p1 = sheet21.max_column + 1
                        p2 = sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2, sheet0.max_row + 1):
                            sheet21.cell(l, p1).value = sheet0.cell(l, i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')
            elif w14.get() == "Не является":
                for i in range(2, sheet0.max_column + 1):
                    if (entry5.get()).lower() != (sheet0.cell(k, i).value).lower():
                        p1 = sheet21.max_column + 1
                        p2 = sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2, sheet0.max_row + 1):
                            sheet21.cell(l, p1).value = sheet0.cell(l, i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')

            elif w14.get() == "Начинается с":
                for i in range(2, sheet0.max_column + 1):
                    if ((sheet0.cell(k, i).value).lower()).startswith((entry5.get()).lower()):
                        p1 = sheet21.max_column + 1
                        p2 = sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2, sheet0.max_row + 1):
                            sheet21.cell(l, p1).value = sheet0.cell(l, i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')

            elif w14.get() == "Оканчивается на":
                for i in range(2, sheet0.max_column + 1):
                    if ((sheet0.cell(k, i).value).lower()).endswith((entry5.get()).lower()):
                        p1 = sheet21.max_column + 1
                        p2 = sheet21.max_column + 2
                        sheet21.cell(1, p1).value = os.path.basename(join(dir, file)[:-13])
                        for l in range(2, sheet0.max_row + 1):
                            sheet21.cell(l, p1).value = sheet0.cell(l, i).value
                            sheet21.cell(l, p2).value = sheet0.cell(l, 1).value
                            w1.save(file_name1 + '/буфер1.xlsx')
    x1=[]
    for i in range(2,sheet21.max_column,2):
        x1.append(sheet21.cell(2,i).value)
        lbox.insert(END,f"{sheet21.cell(2,i).value}({sheet21.cell(1,i).value})")
    w15.destroy()
    w15 = Combobox(tab5, values=x1)
    w15.place(x=140, y=120, height=25, width=106)
global y
y=0
def button15():
    global y
    k=140
    if y>0:
        for j in range(2, sheet21.max_row + 1):
            globals()['label' + str(j)].destroy()
    for i in range(2, sheet21.max_column+1, 2):
        if w15.get()==sheet21.cell(2,i).value:
            for j in range(2,sheet21.max_row+1):
                globals()['label' + str(j)] = Label(tab5, text=f"{sheet21.cell(j, i+1).value}: {sheet21.cell(j, i).value}", font=("Arial Bold", 8))
                globals()['label' + str(j)].place(x=140, y=k)
                k+=30
    y+=1
# -------------------------Создание окна---------------------------
window = tk.Tk()
window.bind_all("<Key>", _onKeyRelease, "+")
name = extract_text()
c = []
c1 = []
window.geometry('700x400+500+250')
window.title(f"{name}")
tab_control = ttk.Notebook(window)

choices1 = []
choices2 = {}
k = 0

# ---------------------------вкладка_1---------------------------------------
tab1 = ttk.Frame(tab_control)
tab_control.add(tab1, text='Классы')

w1 = Combobox(tab1, values=choices1)
w1.place(relx=0.35, y=70, height=25, relwidth=0.15)
w3 = Combobox(tab1, values=c)
w3.place(relx=0.19, y=70, height=25, relwidth=0.15)

message1 = StringVar()
message2 = StringVar()

canvas = tk.Canvas(tab1, height=400, width=700)

canvas.place(relx=0.66 , y=16 , relheight=0.8, relwidth=0.6)


b = openpyxl.Workbook()
b.save('буфер.xlsx')
b = load_workbook(filename='буфер.xlsx')
sheet = b.active

label = Label(tab1, text="Тип класса:", font=("Arial Bold", 10))
label.place(x=1, y=1)
label = Label(tab1, text="Имя класса:", font=("Arial Bold", 10))
label.place(relx=0.19, y=1)
label = Label(tab1, text="Пример:", font=("Arial Bold", 10))
label.place(relx=0.67, y=1)

entry1 = Entry(tab1, textvariable=message1)
entry1.place(relx=0.19, y=20, height=25, relwidth=0.15)

entry2 = Entry(tab1, textvariable=message2)
entry2.place(x=0, y=70, height=25, relwidth=0.18)

choices = ['abstract', 'concrete']
variable = StringVar(tab1)

w = Combobox(tab1, values=choices)
w.place(x=0, y=20, height=25, relwidth=0.18)

butt = Button(tab1, text='создать класс', bg='lightgreen', command=button1Callback)
butt.place(relx=0.35, y=19, height=25, width=106)
butt = Button(tab1, text='создать подкласс', bg='lightgreen', command=button2Callback)
butt.place(relx=0.51, y=70, height=25, width=106)
butt = Button(tab1, text='ok', bg='lightgreen', command=button3Callback)
butt.place(relx=0.29, y=98, height=25, relwidth=0.05)

label = Label(tab1, text="Имя подкласса:", font=("Arial Bold", 10))
label.place(x=1, y=48)
label = Label(tab1, text="Уровень, на котором\nнаходится:", font=("Arial", 7), justify=LEFT)
label.place(relx=0.19, y=47, height=20)
label = Label(tab1, text="Класс или подкласс,\nк которому принадлежит:", font=("Arial", 7), justify=LEFT)
label.place(relx=0.35, y=45, height=25)

# ---------------------------вкладка_2---------------------------------------
tab2 = ttk.Frame(tab_control)
tab_control.add(tab2, text='Слоты')
tab_control.pack(expand=1, fill='both')
c2=["integer","string","instance"]
message3 = StringVar()
w41 = Combobox(tab2, values=c2)
w41.place(x=130, y=24, height=25, width=106)
w4 = Combobox(tab2, values=c1)
w4.place(x=240, y=24, height=25, width=106)
w5 = Combobox(tab2, values=choices1)
w5.place(x=350, y=24, height=25, width=106)

entry_slot = Entry(tab2, textvariable=message3)
entry_slot.place(x=0, y=24, height=25, width=125)

butt = Button(tab2, text='ok', bg='lightgreen', command=button4Callback)
butt.place(x=320, y=52, height=25, width=25)
butt = Button(tab2, text='ok', bg='lightgreen', command=button5Callback)
butt.place(x=460, y=24, height=25, width=25)

label = Label(tab2, text="Наименование слота:", font=("Arial Bold", 8))
label.place(x=1, y=1)
label = Label(tab2, text="Тип слота:", font=("Arial", 9), justify=LEFT)
label.place(x=129, y=1, height=20)
label = Label(tab2, text="Уровень, на котором\nнаходится:", font=("Arial", 7), justify=LEFT)
label.place(x=239, y=1, height=20)
label = Label(tab2, text="Класс или подкласс,\nк которому принадлежит:", font=("Arial", 7), justify=LEFT)
label.place(x=350, y=1, height=21)

# -------------------------------------вкладка_3---------------------------------------
tab3 = ttk.Frame(tab_control)
tab_control.add(tab3, text='формы')
tab_control.pack(expand=2, fill='both')

label = Label(tab3, text="Уровень, на котором\nнаходится:", font=("Arial Bold", 7), justify=LEFT)
label.place(x=0, y=3)
label = Label(tab3, text="Наименование класса\nили подкласса:", font=("Arial", 7), justify=LEFT)
label.place(x=110, y=8, height=20)
label = Label(tab3, text="Выберете форму:", font=("Arial",9), justify=LEFT)
label.place(x=220, y=4, height=20)

butt = Button(tab3, text='ok', bg='lightgreen', command=button6Callback)
butt.place(x=80, y=60, height=25, width=25)
butt = Button(tab3, text='ok', bg='lightgreen', command=button7Callback)
butt.place(x=190, y=60, height=25, width=25)
butt = Button(tab3, text='ok', bg='lightgreen', command=button8Callback)
butt.place(x=300, y=60, height=25, width=25)

w6 = Combobox(tab3, values=c1)
w6.place(x=0, y=30, height=25, width=106)
w7 = Combobox(tab3, values=choices1)
w7.place(x=110, y=30, height=25, width=106)
w8 = Combobox(tab3, values=choices1)
w8.place(x=220, y=30, height=25, width=106)

# -------------------------------------вкладка_4----------------------------------------
tab4 = ttk.Frame(tab_control)
tab_control.add(tab4, text='Экземпляры класса')
tab_control.pack(expand=3, fill='both')

w9 = Combobox(tab4, values=c1)
w9.place(x=0, y=30, height=25, width=106)
w10 = Combobox(tab4, values=choices1)
w10.place(x=110, y=30, height=25, width=106)

label = Label(tab4, text="Уровень, на котором\nнаходится:", font=("Arial Bold", 7), justify=LEFT)
label.place(x=0, y=6,height=20)
label = Label(tab4, text="Наименование класса\nили подкласса:", font=("Arial", 7), justify=LEFT)
label.place(x=110, y=6, height=20)

butt = Button(tab4, text='ok', bg='lightgreen', command=button9Callback)
butt.place(x=80, y=60, height=25, width=25)
butt = Button(tab4, text='создать экземляр', bg='lightgreen', command=button10Callback)
butt.place(x=220, y=30, height=25, width=106)
butt = Button(tab4, text='Внести данные', bg='lightgreen', command=button11Callback)
butt.place(x=220, y=60, height=25, width=106)

# -------------------------------------вкладка_5----------------------------------------

choices4=["Содержит","Не содержит","Является","Не является","Начинается с","Оканчивается на"]

tab5 = ttk.Frame(tab_control)
tab_control.add(tab5, text='Запросы')
tab_control.pack(expand=4, fill='both')

label = Label(tab5, text="Уровень, на котором\nнаходится:", font=("Arial Bold", 7), justify=LEFT)
label.place(x=0, y=6,height=20)
label = Label(tab5, text="Наименование класса\nили подкласса:", font=("Arial", 7), justify=LEFT)
label.place(x=110, y=6, height=20)
label = Label(tab5, text="Наименование слота:", font=("Arial", 7), justify=LEFT)
label.place(x=220, y=8, height=20)
label = Label(tab5, text="Действие:", font=("Arial", 8), justify=LEFT)
label.place(x=330, y=8, height=20)
label = Label(tab5, text="Фильтр:", font=("Arial", 8), justify=LEFT)
label.place(x=440, y=8, height=20)
label = Label(tab5, text="Список:", font=("Arial", 8), justify=LEFT)
label.place(x=0, y=100, height=20)
label = Label(tab5, text="Подробная информация:", font=("Arial", 7), justify=LEFT)
label.place(x=140, y=100, height=20)

w11 = Combobox(tab5, values=c1)
w11.place(x=0, y=30, height=25, width=106)
w12 = Combobox(tab5, values=choices1)
w12.place(x=110, y=30, height=25, width=106)
w13 = Combobox(tab5, values=choices1)
w13.place(x=220, y=30, height=25, width=106)
w14 = Combobox(tab5, values=choices4)
w14.place(x=330, y=30, height=25, width=106)
w15 = Combobox(tab5, values=choices1)
w15.place(x=140, y=120, height=25, width=106)

entry5 = Entry(tab5, textvariable=message3)
entry5.place(x=440, y=30, height=25, width=106)

butt = Button(tab5, text='ok', bg='lightgreen', command=button12Callback)
butt.place(x=80, y=60, height=25, width=25)
butt = Button(tab5, text='ok', bg='lightgreen', command=button13Callback)
butt.place(x=190, y=60, height=25, width=25)
butt = Button(tab5, text='ok', bg='lightgreen', command=button14Callback)
butt.place(x=520, y=60, height=25, width=25)
butt = Button(tab5, text='ok', bg='lightgreen', command=button15)
butt.place(x=250, y=120, height=25, width=25)

lbox = Listbox(tab5)
lbox.place(x=0, y=120, height=140, width=100)
# ---------------------------------Закрытие программы------------------------------------

window.protocol("WM_DELETE_WINDOW", on_closing)
window.mainloop()