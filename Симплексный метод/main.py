from openpyxl import load_workbook
iu = load_workbook(filename='iu.xlsx')
iu1 = load_workbook(filename='iu1.xlsx')
sheet = iu.active
sheet1=iu1.active
pt0 = 3
pt = sheet.max_row-4
tp  = sheet.max_column-2
z=0
c=[]
def o(sheet):
    c=[]
    for j in range(3, tp+1):
        if sheet.cell(pt, j).value<0:
            c.append(abs(sheet.cell(pt, j).value))
    #print(c)
    for j in range(3, tp+1):
        if abs(sheet.cell(pt, j).value) == max(c):
            h = j
    #print(h) #######################################################################################
    p=[]
    for i in range(3,pt):
        if sheet.cell(i,h).value>0:
            sheet.cell(i,tp+2).value = sheet.cell(i,tp+1).value/sheet.cell(i,h).value
            p.append(sheet.cell(i,tp+2).value)
        else:
            sheet.cell(i, tp + 2).value='-'
    for i in range(3, pt):
        if sheet.cell(i,tp+2).value == min(p):
            k=i
    #print(k) #######################################################################################
    for i in range(3,pt+1):
        for j in range(3,tp+2):
            sheet.cell(i+6, j).value = sheet.cell(i,j).value
    sheet.cell(2, h).value,sheet.cell(k,2).value = sheet.cell(k,2).value,sheet.cell(2, h).value
    sheet.cell(1, h).value, sheet.cell(k, 1).value = sheet.cell(k, 1).value, sheet.cell(1, h).value
    Q=sheet.cell(k, h).value
    print(f'Разрешающий элемент: {Q}')
    for i in range(3,pt+1):
        for j in range(3,tp+2):
            if i!=k:
                sheet.cell(i, j).value=sheet.cell(i+6, j).value-(sheet.cell(k+6, j).value*sheet.cell(i+6, h).value)/Q
    for i in range(3,pt+1): #разреш. столб.
         if sheet.cell(2, h).value=='x5':
             sheet.cell(i, h).value=sheet.cell(i+6, tp+3).value-(sheet.cell(k+6, tp+3).value*sheet.cell(i+6, h).value)/Q
         elif sheet.cell(2, h).value=='x6':
             sheet.cell(i, h).value=sheet.cell(i+6,tp+4).value-(sheet.cell(k+6, tp+4).value*sheet.cell(i+6, h).value)/Q
         elif sheet.cell(2, h).value=='x7':
             sheet.cell(i, h).value=sheet.cell(i+6, tp+5).value-(sheet.cell(k+6, tp+5).value*sheet.cell(i+6, h).value)/Q
    for j in range(3, tp + 2):
            sheet.cell(k, j).value = sheet.cell(k + 6, j).value / Q #строка разреш.
    sheet.cell(k, h).value=1/Q
    iu1.save('iu1.xlsx')
for i in range(3,pt+1):
    for j in range(2,tp+1):
        sheet1.cell(i,j+1).value=sheet.cell(i,j).value#перенос из файла в файл
for j in range(3,tp+1):
    sheet1.cell(1, j).value = abs(sheet1.cell(pt, j).value)
    sheet1.cell(2, j).value = 'x'+str(j-2)
for i in range(3,pt):
    sheet1.cell(i, 1).value = 0
    sheet1.cell(i, 2).value = 'x'+str(i+2)
for j in range(2, tp):
    if sheet.cell(pt, j).value < 0:
        z += 1
l=1
while z!=0:
    z=0
    o(sheet1)
    for j in range(3, tp+1):
        if sheet1.cell(pt, j).value < 0:
            z += 1
    print(f'итерация №{l} \nпродолжить?')
    ol=input()
    l+=1
S=0
for i in range(3,pt):
    S+=(sheet1.cell(i, 1).value*sheet1.cell(i, tp+1).value)
    print(sheet1.cell(i, 1).value,' '*3,sheet1.cell(i, tp+1).value)
print(f'оптимальная цена: {S}')
iu1.save('iu1.xlsx')