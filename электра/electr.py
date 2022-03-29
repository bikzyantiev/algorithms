from openpyxl import load_workbook
import numpy as np
b = load_workbook(filename='iu1.xlsx')
c = load_workbook(filename='iu2.xlsx')
sheet = b.active
my_sheet = c.active
k = 3
Pij = 0
Nij = 0
t = ['B', 'C', 'D', 'E', 'F', 'G']
v=[]
sr=0
pt=sheet.max_row-1
def r(v):
    for m in range(3,pt):
        summa = 0
        for el in v:
            summa += el
        sr=summa/(pt-3)
    return(sr)
print (pt)
for j in range(2,8):
    v = []
    for m in range(3, pt):
        v.append(sheet.cell(m, j).value)
    for i in range(3,pt):
        if j==2 or j==6 or j==7:
            if min(v)<=(sheet.cell(i, j)).value <= ((min(v)+r(v))/2):
                (sheet.cell(i, j)).value=5
            if ((min(v) + r(v)) / 2)<= (sheet.cell(i, j)).value < ((max(v) + r(v)) / 2):
                (sheet.cell(i, j)).value = 10
            if ((max(v) + r(v)) / 2)<= (sheet.cell(i, j)).value <=max(v) :
                (sheet.cell(i, j)).value = 15
        if j==3 or j==4 or j==5:
            if min(v) <= (sheet.cell(i, j)).value < ((min(v) + r(v)) / 2):
                (sheet.cell(i, j)).value = 15
            if ((min(v) + r(v)) / 2) <= (sheet.cell(i, j)).value < ((max(v) + r(v)) / 2):
                (sheet.cell(i, j)).value = 10
            if ((max(v) + r(v)) / 2) <= (sheet.cell(i, j)).value <= max(v):
                (sheet.cell(i, j)).value = 5
b.save('iu1.xlsx')
for i in range(3, pt):
    k += 1
    for j in range(k, pt):
        if i == j:
            (my_sheet.cell(i, j)).value = 'x'
        for o in range(6):
            if sheet[t[o] + str(pt+1)].value == "min":
                if sheet[t[o] + str(i)].value < sheet[t[o] + str(j)].value:
                    Pij = Pij + sheet[t[o] + str(pt)].value
                elif sheet[t[o] + str(i)].value > sheet[t[o] + str(j)].value:
                    Nij = Nij + sheet[t[o] + str(pt)].value
            if sheet[t[o] + str(pt+1)].value == "max":
                if sheet[t[o] + str(i)].value > sheet[t[o] + str(j)].value:
                    Pij = Pij + sheet[t[o] + str(pt)].value
                elif sheet[t[o] + str(i)].value < sheet[t[o] + str(j)].value:
                    Nij = Nij + sheet[t[o] + str(pt)].value
        if Nij != 0:
            Dij = Pij / Nij
            if Pij == 0:
                Dij = "o~o"
        else:
            Dij = "--"
        if Pij != 0:
            Dji = Nij / Pij
            if Nij == 0:
                Dji = "o~o"
        else:
            Dji = "--"
        print(f"D({i-2},{j-2})={Pij}/{Nij}={Dij}")
        print(f"D({j-2},{i-2})={Nij}/{Pij}={Dji}\n")
        if Dij!="--":
            if Dij!="o~o" and Dij>1:
                (my_sheet.cell(i - 1, j - 1)).value = Dij
            elif Dij == "o~o":
                (my_sheet.cell(i - 1, j - 1)).value = Dij
            else:
                (my_sheet.cell(i - 1, j - 1)).value = "--"
        else:
            (my_sheet.cell(i - 1, j - 1)).value = Dij
        if Dji!="--":
            if Dji != "o~o" and Dji > 1:
                (my_sheet.cell(j - 1, i - 1)).value = Dji
            elif Dji == "o~o":
                (my_sheet.cell(j - 1, i - 1)).value = Dji
            else:
                (my_sheet.cell(j - 1, i - 1)).value = "--"
        else:
            (my_sheet.cell(j - 1, i - 1)).value = Dji
        Pij = 0
        Nij = 0
t=float(input('введите порог: '))
for i in range(3, pt):
    for j in range(3, pt):
        if i == j:
            (my_sheet.cell(i-1, j-1)).value = 'x'
for i in range(2,pt-1):
    for j in range(2,pt-1):
        if (my_sheet.cell(i, j)).value != "--" and (my_sheet.cell(i, j)).value != "x" and (my_sheet.cell(i, j)).value != "o~o":
            if (my_sheet.cell(i, j)).value <= t:
                (my_sheet.cell(i, j)).value = "--"
y=[]
y1=[]
k=1
for i in range(2,12):
    for j in range(2,12):
        if (my_sheet.cell(i,j)).value != "--" and (my_sheet.cell(i,j)).value != "x":
            y.append(j-1)
        if (my_sheet.cell(j,i)).value != "--" and (my_sheet.cell(j,i)).value != "x":
            y1.append(j-1)
    if len(y)!=0 or (len(y)==0 and len(y1)!=0):
        my_sheet.cell(k+12, 1).value = (my_sheet.cell(i,1)).value
        my_sheet.cell(k+12, 2).value = len(y)
        my_sheet.cell(k+12, 3).value = len(y1)
        k=k+1
    y=[]
    y1=[]
k=k-1
for j in range(13,k+12):
    for i in range(13,k+12):
        if my_sheet.cell(i, 2).value <= my_sheet.cell(i+1, 2).value:
            t1 = my_sheet.cell(i, 1).value
            my_sheet.cell(i, 1).value=my_sheet.cell(i+1, 1).value
            my_sheet.cell(i + 1, 1).value=t1
            t2 = my_sheet.cell(i, 2).value
            my_sheet.cell(i, 2).value = my_sheet.cell(i + 1, 2).value
            my_sheet.cell(i + 1, 2).value = t2
            t3 = my_sheet.cell(i, 3).value
            my_sheet.cell(i, 3).value = my_sheet.cell(i + 1, 3).value
            my_sheet.cell(i + 1, 3).value = t3
for j in range(13,k+12):
    for i in range(13,k+12):
        if my_sheet.cell(i, 3).value <= my_sheet.cell(i+1, 3).value and my_sheet.cell(i, 2).value == my_sheet.cell(i+1, 2).value:
            t1 = my_sheet.cell(i, 1).value
            my_sheet.cell(i, 1).value = my_sheet.cell(i + 1, 1).value
            my_sheet.cell(i + 1, 1).value = t1
            t2 = my_sheet.cell(i, 2).value
            my_sheet.cell(i, 2).value = my_sheet.cell(i + 1, 2).value
            my_sheet.cell(i + 1, 2).value = t2
            t3 = my_sheet.cell(i, 3).value
            my_sheet.cell(i, 3).value = my_sheet.cell(i + 1, 3).value
            my_sheet.cell(i + 1, 3).value = t3
c.save('iu2.xlsx')
b.save('iu1.xlsx')